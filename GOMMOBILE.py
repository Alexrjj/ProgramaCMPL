# -*- coding: utf-8 -*-
import openpyxl
import time
import shutil
from selenium.common.exceptions import NoSuchElementException, TimeoutException
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By

#  Acessa os dados de login fora do script, salvo numa planilha existente, para proteger as informações de credenciais
dados = openpyxl.load_workbook('C:\\gomnet.xlsx')
login = dados['Plan1']
url = 'http://gomnet.ampla.com/'
gommobile = 'http://gomnet.ampla.com/Mobile/telaPrincipal.aspx'
username = login['A1'].value
password = login['A2'].value
mobUser = login['B1'].value
mobPass = login['B2'].value
wb = openpyxl.load_workbook('ExecucaoGommobile.xlsm')
wb1 = openpyxl.load_workbook('ExecucaoGommobile.xlsm')

driver = webdriver.Chrome()

if __name__ == '__main__':
    driver.get(url)
    # Faz login no sistema GOMNET
    uname = driver.find_element_by_name('txtBoxLogin')
    uname.send_keys(username)
    passw = driver.find_element_by_name('txtBoxSenha')
    passw.send_keys(password)
    submit_button = driver.find_element_by_id('ImageButton_Login').click()
    driver.get(gommobile)

    # Faz login no sistema GOMMOBILE
    mobLogin = driver.find_element_by_id('ctl00_ContentPlaceHolder1_txtLogin')
    mobLogin.send_keys(mobUser)
    mobSenha = driver.find_element_by_id('ctl00_ContentPlaceHolder1_txtSenha')
    mobSenha.send_keys(mobPass)
    entrarBtn = driver.find_element_by_id('ctl00_ContentPlaceHolder1_btnEnvia').click()

    # Sincroniza as tarefas
    WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.ID, 'ctl00_ContentPlaceHolder1_lbSincronizar'))).click()
    okBtn = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[2]/div[3]/div/button/span'))).click()
    # Clica nas tarefas pendentes
    WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.ID, 'ctl00_ContentPlaceHolder1_lbTarefaPendente'))).click()

    # Acessa os dados na planilha 'sobs.xlsx' para começar a trabalhar.
    for sheet in wb.worksheets:
        try:
            # Busca o valor da SOB e clica
            WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.XPATH, "*//tr/td[contains(text(), '" + str(sheet['A2'].value) + "')]"))).click()
            # Pressina o TAB uma vez e depois ENTER, para abrir a janela de inserção de dados para a SOB.
            webdriver.ActionChains(driver).send_keys(Keys.TAB).perform()
            webdriver.ActionChains(driver).send_keys(Keys.RETURN).perform()
            # Seleciona todos os baremos executados
            WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.XPATH, '//*[@id="gvItens"]/tbody/tr[1]/th[1]'))).click()
            # Clica na aba "Geral"
            WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.XPATH, '//*[@id="tabGeral"]'))).click()

            # Identifica a data e hora do campo "Fim Tarefa" e salva numa variável
            dataHora = driver.find_element_by_id('txtFimTarefa').get_attribute('value')

            # Clica na caixa de informação de "Saída"
            infSaida = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.XPATH, '//*[@id="txtDtSaida"]')))
            infSaida.send_keys(dataHora)

            # Clica na caixa de informação de "Início Tarefa"
            iniTarefa = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.XPATH, '//*[@id="txtIniTarefa"]')))
            iniTarefa.send_keys(dataHora)

            # Verifica se a célula B1 da planilha 'sobs.xlsx' consta um 'X' para energizar a SOB. Caso não tenha, finaliza parcialmente.
            try:
                if str(sheet['I2'].value) == 'X' or str(sheet['I2'].value) == 'x':
                    energiza = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.XPATH, '//*[@id="txtEnergizacao"]')))
                    energiza.send_keys(dataHora)
                    iniTarefa.click()
            except NoSuchElementException:
                continue

            # Finaliza a SOB
            m = 0
            while m <= 2:
                webdriver.ActionChains(driver).send_keys(Keys.TAB).perform()
                m += 1
            webdriver.ActionChains(driver).send_keys(Keys.SPACE).perform()
            # Clica no botão "Ok"
            time.sleep(20)
            driver.save_screenshot(str(sheet['A2'].value) + '.png')
            print(str(sheet['A2'].value) + ' finalizada com êxito.')
            webdriver.ActionChains(driver).send_keys(Keys.TAB).perform()
            webdriver.ActionChains(driver).send_keys(Keys.SPACE).perform()
        except TimeoutException:  # Caso não encontre a SOB, abre o arquivo txt e registra o número da SOB não movimentada.
            log = open("GommobileOutput.txt", "a")
            log.write(str(sheet['A2'].value) + "\n")
            log.close()
            continue
